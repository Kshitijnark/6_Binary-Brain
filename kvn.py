
# import openpyxl
# import os 
# wb = openpyxl.load_workbook('testing_rasa.xlsx') #give the full path of the file here
# sh = wb.active

# c1 = sh.cell(row=1, column=3)
# c2 = sh.cell(row=2, column=3)
# c3 = sh.cell(row=3, column=3)
# c4 = sh.cell(row=4, column=3)
# c5 = sh.cell(row=5, column=3)
# mean=(c1+c2+c3+c4+c5)/5
# avg_bpm=90
# if (mean>avg_bpm):




import openpyxl
from time import *
import os
# Give the location of the file 
#path = 'testing_rasa.xlsx'
# To open the workbook 
# workbook object is created 
wb_obj = openpyxl.load_workbook(r'D:\SIP_Team_C\\6_Binary-Brain\\testing_rasa.xlsx')
# Get workbook active sheet object 
# from the active attribute 
sheet_obj = wb_obj.active
sum=0
for i in range(1,6):
    cell_obj = sheet_obj.cell(row = i, column = 3)
    sum=sum+cell_obj.value
average=sum/5
print(average)
bpm_range=80
if average>bpm_range:
    os.system('start cmd /k start1\\stress1.cmd')
    sleep(5)
    os.system('start cmd /k start1\\stress2.cmd')
    sleep(15)
    os.system('start cmd /k start1\\stress3.cmd')